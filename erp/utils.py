from openpyxl import Workbook
import os
from django.conf import settings
from datetime import datetime


def export_document_to_excel(document):
    wb = Workbook()
    ws = wb.active

    title_map = {
        "PR": "Purchase Requisition",
        "RFQ": "Request For Quotation",
        "PO": "Purchase Order",
    }

    ws.title = title_map.get(document.doc_type, "Document")

    ws["A1"] = title_map.get(document.doc_type, "Document")

    ws["A3"] = "Company"
    ws["B3"] = document.company.name if document.company else ""

    ws["A4"] = "Partner"
    ws["B4"] = document.partner.name if document.partner else ""

    ws["A5"] = "Document No"
    ws["B5"] = document.doc_no or ""

    ws["A7"] = "No"
    ws["B7"] = "Product"
    ws["C7"] = "Specification"
    ws["D7"] = "Unit"
    ws["E7"] = "Quantity"
    ws["F7"] = "Price"
    ws["G7"] = "Amount"

    row = 8
    total = 0

    for item in document.items.all():
        price = item.price or 0
        qty = item.qty or 0
        amount = qty * price

        ws[f"A{row}"] = item.row_no or (row - 7)
        ws[f"B{row}"] = item.product.name if item.product else item.name
        ws[f"C{row}"] = item.spec
        ws[f"D{row}"] = item.unit
        ws[f"E{row}"] = qty
        ws[f"F{row}"] = price
        ws[f"G{row}"] = amount

        total += amount
        row += 1

    ws[f"F{row + 1}"] = "Total"
    ws[f"G{row + 1}"] = total

    # 파일명
    file_name = f"{document.doc_no}.xlsx" if document.doc_no else f"{document.id}.xlsx"

    file_path = os.path.join(settings.MEDIA_ROOT, file_name)

    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
    wb.save(file_path)

    return file_name
