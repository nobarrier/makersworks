from django.shortcuts import render

# Create your views here.
from django.http import HttpResponse
from .models import Document
from .utils import export_document_to_excel
from django.shortcuts import redirect
from .models import Product, Company, Partner
import openpyxl
from django.shortcuts import render, redirect
from .models import Product, BOM, Company, Partner, Document, DocumentItem
from .models import BOM
from .models import BOMGroup
from .models import BOMLine
from openpyxl.utils import get_column_letter

import os
from django.conf import settings


def export_po_view(request, doc_id):
    document = Document.objects.get(id=doc_id)

    file_name = export_document_to_excel(document)
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)

    with open(file_path, "rb") as f:
        response = HttpResponse(
            f.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename={file_name}"
        return response


from django.shortcuts import render


def po_list_view(request):
    documents = Document.objects.all()
    return render(request, "erp/po_list.html", {"documents": documents})


def create_po_view(request):
    product = Product.objects.first()
    company = Company.objects.get(name="Innoascend")
    partner = Partner.objects.get(name="JPARTS")

    product.create_purchase_order(company, partner, 5)

    return redirect("/po-list/")


def bom_input_view(request):
    return render(request, "erp/bom_input.html")


def upload_bom_view(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            return redirect("/upload-bom/")

        import openpyxl

        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        excel_html = sheet_to_html(ws)

        # 🔥 BOM_ID (엑셀 기준 B2)
        bom_id = str(ws["B2"].value).strip() if ws["B2"].value else file.name

        # 🔥 그룹 생성 또는 가져오기
        group, created = BOMGroup.objects.get_or_create(name=bom_id)

        # 🔥 기존 데이터 삭제 (덮어쓰기)
        if not created:
            BOMLine.objects.filter(group=group).delete()

        excel_rows = []
        created_count = 0

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row:
                continue

            row = list(row) + [None] * 30

            # 🔥 엑셀 원본 (비교용)
            excel_row = {
                "row_no": row[0],
                "mpn": row[1],
                "category": row[2],
                "maker": row[3],
                "item_name": row[4],
                "spec": row[5],
                "unit": row[6],
                "qty": row[7],
                "unit_price": row[8],
                "amount": row[9],
                "supply_amount": row[10],
                "supplier": row[11],
                "draft": row[12],
                "rfq": row[13],
                "quoted": row[14],
                "confirmed": row[15],
                "ordered": row[16],
                "delivered": row[17],
                "remark": row[18],
            }
            excel_rows.append(excel_row)

            # 🔥 번호 기준 필터
            if not row[0]:
                continue

            # 🔥 DB 저장 (엑셀 1:1 매핑)
            BOMLine.objects.create(
                group=group,
                row_no=row[0],
                mpn=str(row[1]).strip() if row[1] else "",
                category=str(row[2]).strip() if row[2] else "",
                maker=str(row[3]).strip() if row[3] else "",
                item_name=str(row[4]).strip() if row[4] else "",
                spec=str(row[5]).strip() if row[5] else "",
                unit=str(row[6]).strip() if row[6] else "",
                qty=float(row[7]) if isinstance(row[7], (int, float)) else None,
                unit_price=float(row[8]) if isinstance(row[8], (int, float)) else None,
                amount=float(row[9]) if isinstance(row[9], (int, float)) else None,
                supply_amount=(
                    float(row[10]) if isinstance(row[10], (int, float)) else None
                ),
                supplier=str(row[11]).strip() if row[11] else "",
                draft=bool(row[12]),
                rfq=bool(row[13]),
                quoted=bool(row[14]),
                confirmed=bool(row[15]),
                ordered=bool(row[16]),
                delivered=bool(row[17]),
                remark=str(row[18]).strip() if row[18] else "",
            )

            created_count += 1

        lines = BOMLine.objects.filter(group=group).order_by("row_no")

        return render(
            request,
            "erp/bom_list.html",
            {
                "excel_html": excel_html,
                "excel_rows": excel_rows,
                "lines": lines,
                "bom_id": group.name,
                "group": group,  # 🔥 이거 추가
            },
        )

    return render(request, "erp/upload_bom.html")


def bom_list_view(request, group_id):
    lines = BOMLine.objects.filter(group_id=group_id).order_by("row_no")
    return render(request, "erp/bom_list.html", {"lines": lines})


def sheet_to_html(ws):
    html = "<table border='1' style='border-collapse:collapse;'>"

    for row in ws.iter_rows(values_only=True):
        html += "<tr>"
        for cell in row:
            value = "" if cell is None else str(cell)
            html += f"<td style='padding:4px;'>{value}</td>"
        html += "</tr>"

    html += "</table>"
    return html


def delete_bom_group(request, group_id):
    BOMLine.objects.filter(group_id=group_id).delete()
    BOMGroup.objects.filter(id=group_id).delete()
    return redirect("/admin/erp/bomgroup/")


from erp.erputils import create_pr_from_bom


def create_pr_from_bom_view(request, group_id):
    company, _ = Company.objects.get_or_create(name="Innoascend")
    group = BOMGroup.objects.get(id=group_id)

    document = create_pr_from_bom(group, company)

    return redirect(f"/export-po/{document.id}/")


from erp.erp_docs.outgoing.rfq import generate_rfq_from_pr
from erp.erp_docs.outgoing.po import generate_po


def create_rfq_from_bom_view(request, group_id):
    company, _ = Company.objects.get_or_create(name="Innoascend")
    group = BOMGroup.objects.get(id=group_id)

    # 기존 PR 생성
    document = create_pr_from_bom(group, company)

    # RFQ 생성
    rfq_doc = generate_rfq_from_pr(document)

    return redirect(f"/export-po/{rfq_doc.id}/")


def create_po_from_bom_view(request, group_id):
    company, _ = Company.objects.get_or_create(name="Innoascend")
    group = BOMGroup.objects.get(id=group_id)

    document = create_pr_from_bom(group, company)

    po_doc = generate_po(document)

    return redirect(f"/export-po/{po_doc.id}/")
